from lxml import etree
from Spider.Seleium.msg_xpath import dialog
from Spider.Seleium.tool import unit_tool
from Spider.Seleium.tool import save
import openpyxl as op

# 多线程爬虫
from queue import Queue


def setup():
    # 打开保存文件
    wb = op.load_workbook('C:\\Users\\30309\\Desktop\\GovSpider\\Spider\\Seleium\\data.xlsx')

    # print(wb.get_named_ranges())  # 输出工作页索引范围
    # print(wb.get_sheet_names())  # 输出所有工作页的名称
    # 取第一张表
    # ws = wb.get_sheet_names()
    # table = wb.get_sheet_by_name(ws[0])
    table = wb.active
    # print(table.title)  # 输出表名
    nrows = table.max_row  # 获得行数
    ncol = table.max_column  # 获得行数
    print(nrows, ncol)

    # 读取配置文件
    head, endLine = save.readLog()

    # 从配置文件中获取是否添加头部信息
    have_head = False
    if head.find('True') != -1:
        have_head = True

    dialog_flag = False
    prevLine = []
    end = False
    # 爬取第1到第25页数据
    for page in range(1, 25):
        root = f'http://www.ccgp.gov.cn/cggg/zygg/gkzb/index_{page}.htm'
        html = etree.HTML(unit_tool.get_html(root))
        paths = html.xpath(dialog.href)
        for path in paths:
            print(dialog.page_root + path)
            page_msg = unit_tool.get_html(dialog.page_root + path)

            list_bs4 = unit_tool.bs4Msg(page_msg)
            if len(list_bs4) == 0:
                continue

            bs4 = list_bs4[0]
            msg = unit_tool.bs4ToStr(bs4.select('td'))
            have_file = unit_tool.haveFile(msg)
            head, val = unit_tool.splitList(msg)
            unit_tool.findCity(val)
            if have_file:
                file_id = unit_tool.getFile(bs4)
                unit_tool.insertFile(file_id, val)
                # print(file_id)
                for sub_url in file_id:
                    file_url = dialog.download_root + sub_url
                    # print(file_url)
            if not have_head:
                # 所有文件写入头,头只写一次
                save.InsertExcel(table, head, nrows)
                nrows += 1
                have_head = True
            # 写入数据
            if endLine.find(str(val)) != -1:
                end = True
                break
            if not dialog_flag:
                prevLine = val
                dialog_flag = True
            save.InsertExcel(table, val, nrows)
            nrows += 1
        if end:
            # prevLine = ':'.join(endLine.split(':')[1:])
            print("Dialog: 今日数据爬取完毕")
            break

    # print("DEBUG获取完成")

    # 写入配置文件
    save.writeLog(have_head, prevLine)

    wb.save('data.xlsx')
